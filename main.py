
from fastapi import FastAPI, UploadFile, File, Query, Body
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request
import pandas as pd
import sqlite3
from pathlib import Path
from datetime import datetime, date, time, timedelta
import io
from typing import Optional, Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / "app.sqlite3"

SLOT_MINUTES = 30

GREEN = PatternFill("solid", fgColor="55B4C7")  # requested accent
GRAY = PatternFill("solid", fgColor="E7E6E6")
BLUE = PatternFill("solid", fgColor="BDD7EE")   # for fact

def slot_labels() -> List[str]:
    labels=[]
    t=datetime.combine(date.today(), time(0,0))
    for i in range(int(24*60/SLOT_MINUTES)):
        labels.append(t.strftime("%H:%M"))
        t += timedelta(minutes=SLOT_MINUTES)
    return labels

SLOTS = slot_labels()

def init_db():
    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            raw_json TEXT NOT NULL
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS request_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            request_id INTEGER NOT NULL,
            row_json TEXT NOT NULL,
            FOREIGN KEY(request_id) REFERENCES requests(id)
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS marks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            day TEXT NOT NULL,
            vehicle_plate TEXT NOT NULL,
            kind TEXT NOT NULL,              -- 'schedule' or 'fact'
            slot TEXT NOT NULL,              -- 'HH:MM'
            value INTEGER NOT NULL,          -- 0/1
            UNIQUE(day, vehicle_plate, kind, slot)
        );
        """)

        # справочник ТС
        cur.execute("""
        CREATE TABLE IF NOT EXISTS vehicle_ref (
            vehicle_plate TEXT PRIMARY KEY,
            schedule_text TEXT NOT NULL DEFAULT '',
            regime_start TEXT NOT NULL DEFAULT '',
            regime_end   TEXT NOT NULL DEFAULT ''
        );
        """)

        # справочник палитры (настраиваемые цвета интерфейса)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS palette (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            accent TEXT NOT NULL DEFAULT '55b4c7',
            schedule_fill TEXT NOT NULL DEFAULT '55b4c7',
            plan_fill TEXT NOT NULL DEFAULT '55b4c7',
            fact_fill TEXT NOT NULL DEFAULT '2563eb',
            border TEXT NOT NULL DEFAULT '0b0f14'
        );
        """)
        # ensure singleton row
        cur.execute("INSERT OR IGNORE INTO palette(id) VALUES (1);")
        con.commit()

init_db()

app = FastAPI()
templates = Jinja2Templates(directory=str(APP_DIR / "templates"))
app.mount("/static", StaticFiles(directory=str(APP_DIR / "static")), name="static")

@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/directory", response_class=HTMLResponse)
def directory_page(request: Request):
    return templates.TemplateResponse("directory.html", {"request": request})


@app.get("/palette", response_class=HTMLResponse)
def palette_page(request: Request):
    return templates.TemplateResponse("palette.html", {"request": request})


def _normalize_hex_color(v: Optional[str], default: str) -> str:
    """Return css hex color (#rrggbb)."""
    if not v:
        return default
    s = str(v).strip().lower()
    if not s:
        return default
    if not s.startswith("#"):
        s = "#" + s
    if len(s) == 4 and all(c in "0123456789abcdef" for c in s[1:]):
        s = "#" + s[1]*2 + s[2]*2 + s[3]*2
    if len(s) == 7 and all(c in "0123456789abcdef" for c in s[1:]):
        return s
    return default


def get_palette(con) -> Dict[str, str]:
    cur = con.cursor()
    cur.execute("SELECT accent, schedule_fill, plan_fill, fact_fill, border FROM palette WHERE id=1")
    row = cur.fetchone()
    accent = _normalize_hex_color(row[0] if row else None, "#55b4c7")
    schedule_fill = _normalize_hex_color(row[1] if row else None, "#55b4c7")
    plan_fill = _normalize_hex_color(row[2] if row else None, "#55b4c7")
    fact_fill = _normalize_hex_color(row[3] if row else None, "#2563eb")
    border = _normalize_hex_color(row[4] if row else None, "#0b0f14")
    return {
        "accent": accent,
        "schedule_fill": schedule_fill,
        "plan_fill": plan_fill,
        "fact_fill": fact_fill,
        "border": border,
    }


@app.get("/api/palette")
def api_palette_get():
    with sqlite3.connect(DB_PATH) as con:
        return get_palette(con)


@app.post("/api/palette")
def api_palette_save(payload: Dict[str, Any] = Body(...)):
    accent = _normalize_hex_color(payload.get("accent"), "#55b4c7")
    schedule_fill = _normalize_hex_color(payload.get("schedule_fill"), "#55b4c7")
    plan_fill = _normalize_hex_color(payload.get("plan_fill"), "#55b4c7")
    fact_fill = _normalize_hex_color(payload.get("fact_fill"), "#2563eb")
    border = _normalize_hex_color(payload.get("border"), "#0b0f14")
    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        cur.execute(
            "UPDATE palette SET accent=?, schedule_fill=?, plan_fill=?, fact_fill=?, border=? WHERE id=1",
            (accent.lstrip("#"), schedule_fill.lstrip("#"), plan_fill.lstrip("#"), fact_fill.lstrip("#"), border.lstrip("#")),
        )
        con.commit()
        return get_palette(con)


@app.post("/api/palette/reset")
def api_palette_reset():
    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        cur.execute(
            "UPDATE palette SET accent='55b4c7', schedule_fill='55b4c7', plan_fill='55b4c7', fact_fill='2563eb', border='0b0f14' WHERE id=1"
        )
        con.commit()
        return get_palette(con)


def _normalize_time_cell(v) -> Optional[str]:
    if pd.isna(v): 
        return None
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    # excel time may be datetime.time
    try:
        if hasattr(v, "strftime"):
            return v.strftime("%H:%M")
    except Exception:
        pass
    s=str(v).strip()
    # handle "07:00 +03:00"
    s=s.replace("+03:00","").replace("+00:00","").strip()
    # handle "7:00"
    if len(s)==4 and s[1]==":":
        s="0"+s
    # handle "07:00:00"
    if len(s)>=8 and s[2]==":" and s[5]==":":
        s=s[:5]
    if len(s)==5 and s[2]==":":
        return s
    return None

def _normalize_date_cell(v) -> Optional[str]:
    if pd.isna(v): 
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s=str(v).strip()
    # handle '09.02.2026'
    try:
        dt=datetime.strptime(s, "%d.%m.%Y")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    try:
        dt=datetime.strptime(s, "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None

def _first_not_nan(*vals):
    for v in vals:
        if v is None: 
            continue
        try:
            if pd.isna(v):
                continue
        except Exception:
            pass
        s=str(v).strip()
        if s=="" or s.lower()=="nan":
            continue
        return v
    return None

    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s=str(v).strip()
    # handle '09.02.2026'
    try:
        dt=datetime.strptime(s, "%d.%m.%Y")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    try:
        dt=datetime.strptime(s, "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None

def parse_requests_excel(content: bytes) -> List[Dict[str, Any]]:
    # default: first sheet
    df = pd.read_excel(io.BytesIO(content))
    # Convert to list of dict (keep original column names)
    rows = df.to_dict(orient="records")
    return rows


def _distinct_values(rows: List[Dict[str, Any]], cols: List[str], limit: int = 500) -> Dict[str, List[str]]:
    """Build distinct values for UI multi-select filters (stringified).

    We cap the number of unique values per column to keep UI fast.
    """
    out: Dict[str, List[str]] = {}
    for c in cols:
        uniq = set()
        for r in rows:
            v = r.get(c)
            if v is None:
                continue
            try:
                if pd.isna(v):
                    continue
            except Exception:
                pass
            s = str(v).strip()
            if not s or s.lower() == "nan":
                continue
            uniq.add(s)
            if len(uniq) >= limit:
                break
        out[c] = sorted(uniq, key=lambda x: x.lower())
    return out

@app.post("/api/upload")
async def upload(file: UploadFile = File(...)):
    content = await file.read()
    rows = parse_requests_excel(content)
    with sqlite3.connect(DB_PATH) as con:
        cur=con.cursor()
        cur.execute("INSERT INTO requests(raw_json) VALUES (?)", (json_dumps({"filename": file.filename, "uploaded_at": datetime.utcnow().isoformat()}),))
        request_id = cur.lastrowid
        for r in rows:
            cur.execute("INSERT INTO request_rows(request_id, row_json) VALUES (?,?)", (request_id, json_dumps(r)))
        con.commit()
    # return columns for filter UI + date range
    cols = list(rows[0].keys()) if rows else []
    dates = sorted({ _normalize_date_cell(r.get("Дата подачи")) for r in rows if _normalize_date_cell(r.get("Дата подачи"))})
    values = _distinct_values(rows, cols)
    return {"ok": True, "request_id": request_id, "columns": cols, "dates": dates, "values": values}

def json_dumps(obj: Any) -> str:
    import json
    return json.dumps(obj, ensure_ascii=False, default=str)

def json_loads(s: str) -> Any:
    import json
    return json.loads(s)

def latest_request_id(con) -> Optional[int]:
    cur=con.cursor()
    cur.execute("SELECT id FROM requests ORDER BY id DESC LIMIT 1")
    row=cur.fetchone()
    return row[0] if row else None

def load_rows(con, request_id: int) -> List[Dict[str,Any]]:
    cur=con.cursor()
    cur.execute("SELECT row_json FROM request_rows WHERE request_id=?", (request_id,))
    return [json_loads(r[0]) for r in cur.fetchall()]

def apply_filters(rows: List[Dict[str, Any]], filters: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Apply UI filters.

    Frontend sends a mapping column -> list of selected values.
    If the list is empty / missing - filter is not applied (means "all").
    Matching is exact by string after stripping.
    """
    if not filters:
        return rows
    out = []
    for r in rows:
        ok = True
        for k, v in filters.items():
            if v is None:
                continue
            selected = v if isinstance(v, list) else [v]
            selected = [str(x).strip() for x in selected if str(x).strip()]
            if not selected:
                continue
            cell = r.get(k)
            if cell is None:
                ok = False
                break
            s = str(cell).strip()
            # exact match against one of selected values
            if s not in selected:
                ok = False
                break
        if ok:
            out.append(r)
    return out

def compute_plan(rows: List[Dict[str,Any]], day_iso: str) -> Dict[str, Dict[str, list]]:
    """
    Return {plate: {slot_label: [request_numbers...]}} where slot covered by any request for that vehicle on that day.
    Vehicle key: 'Гос номер итогового ТС' if present else 'Гос номер ТС'
    Plan interval: [Время подачи, Время завершения)
    """
    plan: Dict[str, Dict[str, list]] = {}
    for r in rows:
        d = _normalize_date_cell(r.get("Дата подачи"))
        if d != day_iso:
            continue

        plate_val = _first_not_nan(r.get("Гос номер итогового ТС"), r.get("Гос номер ТС"))
        plate = str(plate_val or "").strip()
        if not plate or plate.lower() == "nan":
            continue

        start = _normalize_time_cell(r.get("Время подачи"))
        end = _normalize_time_cell(r.get("Время завершения"))
        if not start or not end:
            continue

        req_no = str(r.get("Номер заявки") or "").strip()

        try:
            st = datetime.strptime(start, "%H:%M")
            en = datetime.strptime(end, "%H:%M")
        except Exception:
            continue

        # handle wrap over midnight
        if en <= st:
            en = en + timedelta(days=1)

        cur = st
        while cur < en:
            lbl = cur.strftime("%H:%M")
            if lbl in SLOTS:
                plan.setdefault(plate, {}).setdefault(lbl, [])
                if req_no and req_no.lower() != "nan":
                    plan[plate][lbl].append(req_no)
            cur += timedelta(minutes=SLOT_MINUTES)

    # dedupe preserving order
    for p in list(plan.keys()):
        for s in list(plan[p].keys()):
            nums = [n for n in plan[p][s] if n]
            seen = set()
            uniq = []
            for n in nums:
                if n in seen:
                    continue
                seen.add(n)
                uniq.append(n)
            if uniq:
                plan[p][s] = uniq
            else:
                del plan[p][s]
        if not plan[p]:
            del plan[p]
    return plan


def read_marks(con, day_iso: str, kind: str) -> Dict[str, Dict[str,int]]:
    cur=con.cursor()
    cur.execute("SELECT vehicle_plate, slot, value FROM marks WHERE day=? AND kind=?", (day_iso, kind))
    out={}
    for plate,slot,value in cur.fetchall():
        out.setdefault(plate,{})[slot]=int(value)
    return out


def read_vehicle_ref(con, plates: List[str]) -> Dict[str, Dict[str, str]]:
    """Return mapping plate -> {schedule_text, regime_start, regime_end}."""
    if not plates:
        return {}
    q = "SELECT vehicle_plate, schedule_text, regime_start, regime_end FROM vehicle_ref WHERE vehicle_plate IN (%s)" % (
        ",".join(["?"] * len(plates))
    )
    cur = con.cursor()
    cur.execute(q, plates)
    out = {}
    for p, sch, rs, re_ in cur.fetchall():
        out[str(p)] = {
            "schedule_text": sch or "",
            "regime_start": rs or "",
            "regime_end": re_ or "",
        }
    return out


def list_vehicle_ref(con) -> List[Dict[str, str]]:
    cur = con.cursor()
    cur.execute("SELECT vehicle_plate, schedule_text, regime_start, regime_end FROM vehicle_ref ORDER BY vehicle_plate")
    return [
        {
            "vehicle_plate": p,
            "schedule_text": sch or "",
            "regime_start": rs or "",
            "regime_end": re_ or "",
        }
        for p, sch, rs, re_ in cur.fetchall()
    ]


def upsert_vehicle_ref(con, plate: str, schedule_text: str, regime_start: str, regime_end: str):
    cur = con.cursor()
    cur.execute(
        """
        INSERT INTO vehicle_ref(vehicle_plate, schedule_text, regime_start, regime_end)
        VALUES (?,?,?,?)
        ON CONFLICT(vehicle_plate) DO UPDATE SET
            schedule_text=excluded.schedule_text,
            regime_start=excluded.regime_start,
            regime_end=excluded.regime_end
        """,
        (plate, schedule_text, regime_start, regime_end),
    )
    con.commit()


def delete_vehicle_ref(con, plate: str):
    cur = con.cursor()
    cur.execute("DELETE FROM vehicle_ref WHERE vehicle_plate=?", (plate,))
    con.commit()

@app.get("/api/meta")
def meta():
    with sqlite3.connect(DB_PATH) as con:
        rid = latest_request_id(con)
        if not rid:
            return {"has_data": False}
        rows = load_rows(con, rid)
    cols=list(rows[0].keys()) if rows else []
    dates = sorted({ _normalize_date_cell(r.get("Дата подачи")) for r in rows if _normalize_date_cell(r.get("Дата подачи"))})
    values = _distinct_values(rows, cols)
    return {"has_data": True, "request_id": rid, "columns": cols, "dates": dates, "values": values}

@app.get("/api/schedule")
def schedule(day: str = Query(..., description="YYYY-MM-DD"),
             filters: str = Query("", description="JSON string mapping column->selected list")):
    with sqlite3.connect(DB_PATH) as con:
        rid = latest_request_id(con)
        if not rid:
            return JSONResponse({"error":"no uploaded data"}, status_code=400)
        rows = load_rows(con, rid)

        import json
        f = json.loads(filters) if filters else {}
        rows_f = apply_filters(rows, f)

        plan = compute_plan(rows_f, day)
        schedule_marks = read_marks(con, day, "schedule")
        fact = read_marks(con, day, "fact")

    # vehicles list from filtered rows (только Итоговое ТС)
    vehicles: Dict[str, Dict[str, Any]] = {}
    for r in rows_f:
        plate_val = _first_not_nan(r.get("Гос номер итогового ТС"), r.get("Гос номер ТС"))
        plate = str(plate_val or "").strip()
        if not plate or plate.lower() == "nan":
            continue
        vehicles.setdefault(
            plate,
            {
                "vehicle_name": str(_first_not_nan(r.get("Итоговое ТС"), r.get("ТС")) or "").strip(),
                "vehicle_plate": plate,
                "vehicle_class": str(_first_not_nan(r.get("Класс итогового ТС"), r.get("Класс назначенного ТС")) or "").strip(),
                "schedule_text": "",
                "regime_start": "",
                "regime_end": "",
            },
        )

    vehicle_list = sorted(vehicles.values(), key=lambda x: x["vehicle_plate"])
    plates = [v["vehicle_plate"] for v in vehicle_list]

    with sqlite3.connect(DB_PATH) as con:
        ref_map = read_vehicle_ref(con, plates)

    for v in vehicle_list:
        ref = ref_map.get(v["vehicle_plate"], {})
        v["schedule_text"] = ref.get("schedule_text", "")
        v["regime_start"] = ref.get("regime_start", "")
        v["regime_end"] = ref.get("regime_end", "")

    # build effective schedule layer: saved marks + defaults from справочника (если есть режим)
    schedule_effective: Dict[str, Dict[str, int]] = {}
    for v in vehicle_list:
        plate = v["vehicle_plate"]
        schedule_effective[plate] = dict(schedule_marks.get(plate, {}))
        rs = (v.get("regime_start") or "").strip()
        re_ = (v.get("regime_end") or "").strip()
        if rs and re_:
            try:
                st = datetime.strptime(rs, "%H:%M")
                en = datetime.strptime(re_, "%H:%M")
                if en <= st:
                    en = en + timedelta(days=1)
                cur = st
                while cur < en:
                    lbl = cur.strftime("%H:%M")
                    if lbl in SLOTS and lbl not in schedule_effective[plate]:
                        schedule_effective[plate][lbl] = 1
                    cur += timedelta(minutes=SLOT_MINUTES)
            except Exception:
                pass

    return {
        "day": day,
        "slots": SLOTS,
        "vehicles": vehicle_list,
        "plan": plan,
        "schedule": schedule_effective,
        "fact": fact,
        "filtered_count": len(rows_f),
        "total_count": len(rows)
    }

@app.post("/api/mark")
def mark(payload: Dict[str,Any] = Body(...)):
    """
    payload: {day:'YYYY-MM-DD', plate:'', kind:'schedule'|'fact', slot:'HH:MM', value:0|1}
    """
    day = payload.get("day")
    plate = str(payload.get("plate") or "").strip()
    kind = payload.get("kind")
    slot = payload.get("slot")
    value = int(payload.get("value") or 0)

    if kind not in ("schedule","fact"):
        return JSONResponse({"error":"invalid kind"}, status_code=400)
    if slot not in SLOTS:
        return JSONResponse({"error":"invalid slot"}, status_code=400)
    if not day or not plate:
        return JSONResponse({"error":"missing day/plate"}, status_code=400)

    with sqlite3.connect(DB_PATH) as con:
        cur=con.cursor()
        cur.execute("""
        INSERT INTO marks(day, vehicle_plate, kind, slot, value)
        VALUES (?,?,?,?,?)
        ON CONFLICT(day, vehicle_plate, kind, slot) DO UPDATE SET value=excluded.value
        """,(day, plate, kind, slot, value))
        con.commit()
    return {"ok": True}


@app.get("/api/directory")
def api_directory_list():
    with sqlite3.connect(DB_PATH) as con:
        return {"items": list_vehicle_ref(con)}


@app.post("/api/directory/upsert")
def api_directory_upsert(payload: Dict[str, Any] = Body(...)):
    plate = str(payload.get("vehicle_plate") or "").strip()
    schedule_text = str(payload.get("schedule_text") or "").strip()
    regime_start = str(payload.get("regime_start") or "").strip()
    regime_end = str(payload.get("regime_end") or "").strip()
    if not plate:
        return JSONResponse({"error": "missing vehicle_plate"}, status_code=400)
    # normalize times to HH:MM if possible
    regime_start = _normalize_time_cell(regime_start) or regime_start
    regime_end = _normalize_time_cell(regime_end) or regime_end
    with sqlite3.connect(DB_PATH) as con:
        upsert_vehicle_ref(con, plate, schedule_text, regime_start, regime_end)
    return {"ok": True}


@app.post("/api/directory/delete")
def api_directory_delete(payload: Dict[str, Any] = Body(...)):
    plate = str(payload.get("vehicle_plate") or "").strip()
    if not plate:
        return JSONResponse({"error": "missing vehicle_plate"}, status_code=400)
    with sqlite3.connect(DB_PATH) as con:
        delete_vehicle_ref(con, plate)
    return {"ok": True}

def build_export_xlsx(day: str, schedule_payload: Dict[str,Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "График"

    ws.cell(1, 1).value = "Дата"
    try:
        ws.cell(1, 2).value = datetime.strptime(day, "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        ws.cell(1, 2).value = day

    headers = [
        "Итоговое ТС",
        "Гос номер итогового ТС",
        "Класс итогового ТС",
        "График работы",
        "Режим работы",
        "",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(2, i)
        c.value = h
        c.font = Font(bold=True)
        c.fill = GRAY
        c.alignment = Alignment(horizontal="center", vertical="center")

    for j, slot in enumerate(SLOTS, start=7):
        c = ws.cell(2, j)
        c.value = slot
        c.font = Font(bold=True)
        c.fill = GRAY
        c.alignment = Alignment(horizontal="center", vertical="center")

    vehicles = schedule_payload["vehicles"]
    plan = schedule_payload["plan"]
    schedule_layer = schedule_payload["schedule"]
    fact = schedule_payload["fact"]

    r = 3
    for v in vehicles:
        plate = v["vehicle_plate"]
        regime = ""
        rs = (v.get("regime_start") or "").strip()
        re_ = (v.get("regime_end") or "").strip()
        if rs and re_:
            regime = f"{rs} - {re_}"

        ws.cell(r, 1).value = v.get("vehicle_name", "")
        ws.cell(r, 2).value = plate
        ws.cell(r, 3).value = v.get("vehicle_class", "")
        ws.cell(r, 4).value = v.get("schedule_text", "")
        ws.cell(r, 5).value = regime

        ws.cell(r, 6).value = "График работы"
        ws.cell(r + 1, 6).value = "План"
        ws.cell(r + 2, 6).value = "Факт"
        for rr in (r, r + 1, r + 2):
            ws.cell(rr, 6).font = Font(bold=True)
            ws.cell(rr, 6).alignment = Alignment(horizontal="left", vertical="center")

        # merge left columns across 3 rows
        for col in range(1, 6):
            ws.merge_cells(start_row=r, start_column=col, end_row=r + 2, end_column=col)
            ws.cell(r, col).alignment = Alignment(vertical="center")

        for j, slot in enumerate(SLOTS, start=7):
            if schedule_layer.get(plate, {}).get(slot, 0) == 1:
                ws.cell(r, j).fill = GREEN

            nums = plan.get(plate, {}).get(slot, [])
            if nums:
                ws.cell(r + 1, j).fill = GREEN
                # write request number(s) in Plan cells
                if len(nums) == 1:
                    ws.cell(r + 1, j).value = str(nums[0])
                else:
                    ws.cell(r + 1, j).value = f"{nums[0]} +{len(nums)-1}"

            if fact.get(plate, {}).get(slot, 0) == 1:
                ws.cell(r + 2, j).fill = BLUE

            for rr in (r, r + 1, r + 2):
                ws.cell(rr, j).alignment = Alignment(horizontal="center")

        r += 3

    widths = {1: 28, 2: 16, 3: 18, 4: 18, 5: 16, 6: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for j in range(7, 7 + len(SLOTS)):
        ws.column_dimensions[get_column_letter(j)].width = 4
    ws.freeze_panes = "G3"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

@app.get("/export")
def export(day: str = Query(...),
           filters: str = Query("")):
    # reuse schedule computation so export is identical to UI state
    payload = schedule(day=day, filters=filters)
    content = build_export_xlsx(day, payload)
    out_path = DATA_DIR / f"export_{day}.xlsx"
    out_path.write_bytes(content)
    return FileResponse(str(out_path), filename=out_path.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
